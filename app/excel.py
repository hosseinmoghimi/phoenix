
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from django.http import HttpResponse
from .persian import PersianCalendar
from datetime import datetime
from openpyxl import load_workbook
from .settings import *
import os
def get_style():
    font = Font(name='Calibri',
                    size=11,
                    bold=False,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')
    fill = PatternFill(fill_type=None,
                    start_color='FFFFFF',
                    end_color='FF000000')
    border = Border(left=Side(border_style='thin',
                            color='000000'),
                    right=Side(border_style='thin',
                            color='000000'),
                    top=Side(border_style='thin',
                            color='000000'),
                    bottom=Side(border_style='thin',
                                color='000000'),
                    diagonal=Side(border_style=None,
                                color='FF000000'),
                    diagonal_direction=0,
                    outline=Side(border_style=None,
                                color='FF000000'),
                    vertical=Side(border_style=None,
                                color='FF000000'),
                    horizontal=Side(border_style=None,
                                color='FF000000')
                )
    alignment=Alignment(horizontal='center',
                        vertical='bottom',
                        text_rotation=0,
                        wrap_text=False,
                        shrink_to_fit=False,
                        indent=0)
    number_format = 'General'
    protection = Protection(locked=True,
                            hidden=False)
    return {'font':font,
        'fill':fill,
        'border':border,
        'alignment':alignment,
        'number_format':number_format,
        'protection':protection,
    }
class ReportWorkBook:
    
    def __init__(self,file_name=None):
        self.sheets=[]
        self.file_name=file_name
        
    # def to_excel2(self):
    #     date=PersianCalendar().from_gregorian(datetime.now())
    #     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
    #     # response.AppendHeader("Content-Type", "application/vnd.ms-excel");
    #     response["Content-disposition"]=f"attachment; filename={self.file_name}-{date}.xlsx"
    #     work_book=Workbook()
    #     for i,sheet in enumerate(self.sheets):
    #         if i==0:
    #             worksheet=work_book.active
    #         else:
    #             worksheet=work_book.create_sheet(f'sheet#{i}')
    #         sheet.get_worksheet(worksheet,blank_sheet=True)
    #     work_book.save(response)
    #     return response
    def to_excel(self):
        date=PersianCalendar().from_gregorian(datetime.now())
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        # response.AppendHeader("Content-Type", "application/vnd.ms-excel");
        response["Content-disposition"]=f"attachment; filename={self.file_name}-{date}.xlsx"
        REPORT_ROOT=os.path.join(BASE_DIR,'report')
        if self.file_name is None:
            work_book = Workbook()
            for sheet in self.sheets:
                work_book.create_sheet(sheet.title)
        else:
            filename =os.path.join(REPORT_ROOT,self.file_name+'.xlsx')        
            work_book = load_workbook(filename = filename)
        for i,sheet in enumerate(self.sheets):
            work_sheet=work_book.worksheets[i]
            sheet.get_worksheet(work_sheet,blank_sheet=False)
        work_book.save(response)
        work_book.close()
        return response
class ReportSheet:
    def __init__(self,data=None,title=None,table_headers=None):
        self.data=data
        if title is None:
                self.title='گزارش'
        else:
            self.title=title
        if table_headers is None:
            self.table_headers=[]
            for key in self.data[0].keys():
                self.table_headers.append(key)
        else:
            self.table_headers=table_headers
    def get_worksheet(self,worksheet,blank_sheet=False):      
        style=get_style()                

        worksheet.title = self.title
        worksheet.sheet_view.rightToLeft = True
        if len(self.data)<1:
            return worksheet
        # Define the titles for columns

        row_num = 1
        cell = worksheet.cell(row=row_num, column=1)
        cell.value = self.title
        cell.border=style['border']
        cell.alignment=style['alignment']
        if blank_sheet:
            column=len( self.table_headers)
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=column)       
        row_num+=1
        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(self.table_headers, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            if blank_sheet:
                cell.alignment=style['alignment']
                cell.border=style['border']

        # Iterate through all movies
        for data_item in self.data:
            row_num += 1
            
            col_num=1
            # Assign the data for each cell of the row 
            for cell_value in data_item:
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = data_item[cell_value]
                cell.border=style['border']
                col_num+=1
        return worksheet
